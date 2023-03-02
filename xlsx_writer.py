import openpyxl

def write_schedule_to_xlsx(people, schedule):
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Write header row with days of week
    header_font = openpyxl.styles.Font(bold=True)
    header_alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.cell(row=1, column=1, value="Days of Week").font = header_font
    for i, day in enumerate(schedule.keys()):
        ws.cell(row=1, column=i+2, value=day).font = header_font
        ws.cell(row=1, column=i+2).alignment = header_alignment

    # Write schedule data
    data_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    for i, person in enumerate(people):
        ws.cell(row=i+2, column=1, value=person).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=i+2, column=1).fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        for j, day in enumerate(schedule.keys()):
            cell = ws.cell(row=i+2, column=j+2)
            if person in schedule[day]:
                cell.value = "X"
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            cell.alignment = data_alignment

    # Adjust column widths to fit content
    for column in ws.columns:
        column_dimensions = column[0].column_letter
        ws.column_dimensions[column_dimensions].auto_size = True

    # Apply some additional formatting to the worksheet
    ws.freeze_panes = "B2"
    ws.sheet_view.zoomScale = 125

    # Save the workbook to a file
    wb.save("schedule.xlsx")
